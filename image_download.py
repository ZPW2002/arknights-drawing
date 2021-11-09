# -*- coding: UTF-8 -*-
import requests
import openpyxl
import image_url
import os
import time

head = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:91.0) Gecko/20100101 Firefox/91.0"}


def get_image(name, url, path_):
    if url:
        path = path_ + '立绘_%s_2.png' % name
        print(path)
        image = open(path, 'wb')
        response = requests.get(url, headers=head)
        image.write(response.content)
    else:
        print('no url')


def img_down(path):

    workbook = openpyxl.load_workbook('name.xlsx')
    sheet = workbook.active
    
    os.mkdir(path)
    os.mkdir(path + '/stage1')
    os.mkdir(path + '/stage2')
    os.mkdir(path + '/skin')

    for i in range(1, image_url.img_num + 1):
        time.sleep(0.5)
        print("%d/%d" % (i, image_url.img_num))

        # 下载初始立绘
        temp_name = sheet.cell(i, 1).value
        temp_url = 'http:' + sheet.cell(i, 2).value
        get_image(temp_name, temp_url, path + '/stage1/')

        # 下载精二立绘
        temp_url = sheet.cell(i, 3).value
        if not temp_url:
            continue
        temp_url = 'http:' + temp_url
        temp_name = sheet.cell(i, 1).value
        get_image(temp_name, temp_url, path + '/stage2/')
        

    # 下载皮肤
    for i in range(1, sheet.max_row + 1):
        print('下载皮肤中...(%d/%d)' % (i, sheet.max_row))
        temp_url = 'http:' + sheet.cell(i, 4).value
        get_image(str(i), temp_url, path + '/skin/')

    print("完成！")
