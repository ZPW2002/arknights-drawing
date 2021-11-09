# -*- coding: UTF-8 -*-
import requests
from bs4 import BeautifulSoup
import re
import openpyxl
import time


head = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:91.0) Gecko/20100101 Firefox/91.0"}
img_num = 214


# URL编码
def encode_(tar):
    tar = str(tar.encode('utf-8')).lstrip("'b").rstrip("'")
    return tar.replace(r'\x', '%')


# 用beautifulsoup缩小范围
def get_soup(name):
    url = 'http://prts.wiki/w/' + encode_(name)
    response = requests.get(url, headers=head)
    soup = BeautifulSoup(response.text.encode('utf-8'), 'html.parser')
    return soup


# 正则匹配url
def get_link(soup, skin):
    item_str = str(soup.find_all('div', id=skin)[0])
    link = re.findall(re.compile('src="(.+)" width'), item_str)
    return link


def get_url():

    print('获取图片url...')
    workbook = openpyxl.load_workbook('name.xlsx')
    sheet = workbook.active

    global img_num
    img_num = sheet.max_row
    print(img_num)

    skin_num = 1
    times = img_num
    i = 0
    while True:
        times -= 1
        i += 1
        if not times:
            break 

        try:
            # for i in range(start, get_name.img_num + 1):
            time.sleep(1.5)

            print("\n%s(%d/%d)" % (sheet.cell(i, 1).value, i, img_num))
            soup = get_soup(sheet.cell(i, 1).value)

            # 初始立绘
            stage0_link = get_link(soup, 'img-stage0')
            sheet.cell(i, 2).value = stage0_link[0]
            print(stage0_link[0])

            # 精二立绘
            stage2_link = get_link(soup, 'stage2_default')
            if stage2_link:
                sheet.cell(i, 3).value = stage2_link[0]
                print(stage2_link[0])
            else:
                print('no stage2')

            # 干员皮肤
            count = 0
            while count < 10:
                count += 1
                skin = 'img-skin' + str(count)
                link = get_link(soup, skin)
                if link:
                    sheet.cell(skin_num, 4).value = link[0]
                    skin_num += 1
                    print(link[0])
                else:
                    break
        
            workbook.save('name.xlsx')

        except ValueError:
            print('重新连接')
            count += 1
            i -= 1
            continue

    print('图片url获取完毕')
    workbook.save('name.xlsx')
